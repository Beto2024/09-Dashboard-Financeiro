from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from models.models import db, Transaction, CATEGORIES
from sqlalchemy import extract
from datetime import datetime, date
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

reports_bp = Blueprint("reports", __name__)


def get_filtered_transactions(month, year, t_type, category):
    query = Transaction.query
    if month:
        query = query.filter(extract("month", Transaction.date) == int(month))
    if year:
        query = query.filter(extract("year", Transaction.date) == int(year))
    if t_type:
        query = query.filter(Transaction.type == t_type)
    if category:
        query = query.filter(Transaction.category == category)
    return query.order_by(Transaction.date.desc()).all()


@reports_bp.route("/reports")
def index():
    today = date.today()
    month = request.args.get("month", "")
    year = request.args.get("year", str(today.year))
    t_type = request.args.get("type", "")
    category = request.args.get("category", "")

    transactions = get_filtered_transactions(month, year, t_type, category)

    total_income = sum(t.amount for t in transactions if t.type == "Receita")
    total_expense = sum(t.amount for t in transactions if t.type == "Despesa")
    balance = total_income - total_expense

    months = [
        (1, "Janeiro"), (2, "Fevereiro"), (3, "Março"), (4, "Abril"),
        (5, "Maio"), (6, "Junho"), (7, "Julho"), (8, "Agosto"),
        (9, "Setembro"), (10, "Outubro"), (11, "Novembro"), (12, "Dezembro"),
    ]
    years = list(range(today.year - 3, today.year + 2))

    return render_template(
        "reports.html",
        transactions=transactions,
        categories=CATEGORIES,
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        selected_month=month,
        selected_year=year,
        selected_type=t_type,
        selected_category=category,
        months=months,
        years=years,
    )


@reports_bp.route("/reports/export/csv")
def export_csv():
    month = request.args.get("month", "")
    year = request.args.get("year", "")
    t_type = request.args.get("type", "")
    category = request.args.get("category", "")

    transactions = get_filtered_transactions(month, year, t_type, category)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Descrição", "Valor", "Data", "Tipo", "Categoria"])
    for t in transactions:
        writer.writerow([t.id, t.description, t.amount, t.date.strftime("%d/%m/%Y"), t.type, t.category])

    output.seek(0)
    byte_output = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    byte_output.seek(0)

    filename = f"transacoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        byte_output,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/reports/export/excel")
def export_excel():
    month = request.args.get("month", "")
    year = request.args.get("year", "")
    t_type = request.args.get("type", "")
    category = request.args.get("category", "")

    transactions = get_filtered_transactions(month, year, t_type, category)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transações"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    headers = ["ID", "Descrição", "Valor (R$)", "Data", "Tipo", "Categoria"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.id)
        ws.cell(row=row, column=2, value=t.description)
        ws.cell(row=row, column=3, value=t.amount)
        ws.cell(row=row, column=4, value=t.date.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=5, value=t.type)
        ws.cell(row=row, column=6, value=t.category)

        if t.type == "Receita":
            fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        else:
            fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill

    col_widths = [6, 35, 15, 15, 12, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Summary sheet
    ws2 = wb.create_sheet("Resumo")
    total_income = sum(t.amount for t in transactions if t.type == "Receita")
    total_expense = sum(t.amount for t in transactions if t.type == "Despesa")
    balance = total_income - total_expense

    ws2.append(["Resumo Financeiro"])
    ws2.append([])
    ws2.append(["Total de Receitas", f"R$ {total_income:,.2f}"])
    ws2.append(["Total de Despesas", f"R$ {total_expense:,.2f}"])
    ws2.append(["Saldo", f"R$ {balance:,.2f}"])
    ws2.append(["Total de Transações", len(transactions)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_financeiro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
